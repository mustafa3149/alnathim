"""
Al-Nathim ISP Management System — Flask Application
===================================================
Phase 10: Architecture Reboot — raw SQLite via database.py (parameterized queries),
single-tenant admin/agent users table, Jinja2 templates preserved as-is.
"""

import io
import logging
import os
import secrets
from datetime import datetime, timedelta, date

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, g, abort

from config import (
    SECRET_KEY,
    LOCAL_DB_PATH,
    SESSION_LIFETIME_MINUTES,
    COOKIE_SECURE,
    RELAY_URL,
)
import database as db
from auth import auth_bp, login_required, admin_required, current_user_id
from billing_system.mikrotik_sync import sync_customer_debt, sync_mikrotik_status
from network_tools.ping import ping_host, validate_host
from snmp_monitor.signal_monitor import SignalMonitor
from mikrotik_api.mikrotik_manager import MikroTikManager

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY
app.config["SESSION_PERMANENT"] = True
app.config["PERMANENT_SESSION_LIFETIME"] = SESSION_LIFETIME_MINUTES * 60
# ── Cookie hardening (Phase 14.5) ─────────────────────────
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = COOKIE_SECURE

app.register_blueprint(auth_bp)

# Initialize DB + seed defaults on startup
db.init_db()


# ──────────────────────────────────────────────
#  CSRF PROTECTION + SECURITY HEADERS (Phase 14.5)
# ──────────────────────────────────────────────
_CSRF_EXEMPT_PATHS = {"/login", "/register", "/logout", "/api/agent/signal", "/api/remote/login"}


def get_csrf_token():
    """Return the session CSRF token, generating one if absent."""
    if not session.get("_csrf_token"):
        session["_csrf_token"] = secrets.token_hex(32)
    return session["_csrf_token"]


@app.before_request
def csrf_protect():
    """Reject unsafe (POST/PUT/PATCH/DELETE) requests without a valid CSRF header.

    Pre-auth endpoints (login/register/logout) are exempt — they are protected
    by the per-IP rate limiter and SameSite cookie instead.
    """
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return None
    if request.path in _CSRF_EXEMPT_PATHS:
        return None
    token = request.headers.get("X-CSRF-Token", "")
    if not token or token != session.get("_csrf_token"):
        # Allow FORM posts to carry the token in a field too.
        if request.form.get("_csrf_token") != session.get("_csrf_token"):
            return jsonify({"ok": False, "error": "طلب غير مصرح به (CSRF)"}), 403
    return None


@app.after_request
def security_headers(response):
    """Add baseline hardening headers to every response."""
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    return response


# ──────────────────────────────────────────────
#  HELPERS & FORMATTERS
# ──────────────────────────────────────────────

DATETIME_DB = "%Y-%m-%d %H:%M"
DATETIME_DB_S = "%Y-%m-%d %H:%M:%S"
ARABIC_DIGITS = str.maketrans("0123456789", "٠١٢٣٤٥٦٧٨٩")


def now_dt():
    """Current datetime with seconds/microseconds zeroed."""
    return datetime.now().replace(second=0, microsecond=0)


def fmt_dt(dt):
    """Format datetime for display: YYYY-MM-DD HH:MM AM/PM (handles str)."""
    if dt is None or dt == "":
        return "—"
    if isinstance(dt, str):
        for fmt in (DATETIME_DB, DATETIME_DB_S):
            try:
                dt = datetime.strptime(dt, fmt)
                break
            except ValueError:
                continue
        else:
            return dt  # fallback: return raw string
    return dt.strftime("%Y-%m-%d %I:%M %p")


def fmt_dt_db(dt):
    """Format datetime for DB storage."""
    if dt is None:
        return None
    if isinstance(dt, str):
        return dt
    return dt.strftime(DATETIME_DB)


def to_arabic_num(s):
    """Convert Western digits to Arabic numerals."""
    return str(s).translate(ARABIC_DIGITS)


def fmt_num(n, style="AR"):
    """Format with thousands separators, optionally Arabic digits."""
    try:
        f = f"{int(round(float(n))):,}"
    except (ValueError, TypeError):
        f = "0"
    return to_arabic_num(f) if style == "AR" else f


def get_numeral_style():
    """Return the current numeral style from settings (AR/EN)."""
    try:
        settings = db.get_settings()
        return settings.get("numeral_style", "AR")
    except Exception:
        return "AR"


def get_gen_info():
    """Return the generator_info row (company profile), or None."""
    return db.get_generator_info()


def _audit(action, target_type="", target_id=None, details=""):
    """Log an action performed by the current session user (who did what)."""
    uid = current_user_id()
    uname = session.get("user_name") or ""
    if uid:
        db.log_action(user_id=uid, username=uname, action=action,
                      target_type=target_type, target_id=target_id, details=details)


def is_admin():
    """Return True if the current session is an admin."""
    return session.get("user_role") == "admin"


def last_day_of_month(year, month):
    """Return the last day of the given month/year (1-31)."""
    if month == 12:
        return (date(year + 1, 1, 1) - timedelta(days=1)).day
    return (date(year, month + 1, 1) - timedelta(days=1)).day


# ──────────────────────────────────────────────
#  JINJA2 FILTERS
# ──────────────────────────────────────────────

@app.template_filter("dinar")
def dinar_filter(value):
    n = getattr(g, "numeral_style", "AR") or "AR"
    if value is None:
        return (to_arabic_num("0") if n == "AR" else "0") + " د.ع"
    try:
        return fmt_num(value, n) + " د.ع"
    except (ValueError, TypeError):
        return (to_arabic_num("0") if n == "AR" else "0") + " د.ع"


@app.template_filter("invoice_no")
def invoice_no_filter(value):
    n = getattr(g, "numeral_style", "AR") or "AR"
    try:
        f = f"#INV-{int(value):04d}"
    except (ValueError, TypeError):
        f = "#INV-0"
    return to_arabic_num(f) if n == "AR" else f


@app.template_filter("numeral")
def numeral_filter(value):
    n = getattr(g, "numeral_style", "AR") or "AR"
    if value is None:
        return to_arabic_num("0") if n == "AR" else "0"
    try:
        return fmt_num(int(round(float(value))), n)
    except (ValueError, TypeError):
        return to_arabic_num("0") if n == "AR" else "0"


@app.template_filter("datetime_display")
def datetime_display_filter(value):
    return fmt_dt(value)


@app.template_filter("package_en")
def package_en_filter(value):
    """Map Arabic package names to English for receipts."""
    mapping = {
        "ايكونومي": "Economy", "بلاس": "Plus", "ستاندرد": "Standard",
        "توربو": "Turbo", "مور": "More", "بزنس برو": "Business Pro",
    }
    if value is None:
        return "—"
    return mapping.get(str(value).strip(), str(value))


@app.context_processor
def inject_globals():
    """Inject into templates: now_str + numeral style + role flags."""
    numeral = get_numeral_style()
    g.numeral_style = numeral
    g.is_admin = is_admin()
    g.is_agent = session.get("user_role") == "agent"
    return {
        "now_str": fmt_dt_db(now_dt()),
        "numeral_style": numeral,
        "is_admin": g.is_admin,
        "is_agent": g.is_agent,
        "csrf_token": get_csrf_token(),
        "relay_url": RELAY_URL,
    }


# ──────────────────────────────────────────────
#  ROUTES: PAGES (login-required)
# ──────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    """Dashboard KPIs."""
    now = now_dt()
    month, year = now.month, now.year
    stats = db.dashboard_stats(month, year)
    collected_today = db.total_payments_today()

    alert_days = request.args.get("alert_days", 3, type=int)
    alert_date = now + timedelta(days=alert_days)

    expiring_rows = db.customers_expiring_between(
        fmt_dt_db(now), fmt_dt_db(alert_date)
    )
    expiring = []
    for e in expiring_rows:
        try:
            rd = datetime.strptime(str(e["renewal_date"]), DATETIME_DB)
        except ValueError:
            try:
                rd = datetime.strptime(str(e["renewal_date"]), DATETIME_DB_S)
            except ValueError:
                rd = now + timedelta(days=alert_days + 1)
        hours_left = max((rd - now).total_seconds() / 3600.0, 0)
        expiring.append({
            "id": e["id"],
            "name": e["name"],
            "phone": e["phone"],
            "package_name": e["package_name"],
            "renewal_date": e["renewal_date"],
            "hours_left": hours_left,
        })

    return render_template(
        "dashboard.html",
        active_customers=stats["active_customers"],
        total_packages=stats["total_packages"],
        expected_income=stats["expected_income"],
        collected=stats["collected"],
        collected_today=collected_today,
        total_debt=stats["total_debt"],
        expiring=expiring,
        alert_days=alert_days,
    )


@app.route("/customers")
@login_required
def customers_page():
    """List customers with filters."""
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "all")
    region_filter = request.args.get("region", "").strip()
    debt_filter = request.args.get("debt", "").strip()
    sort_by = request.args.get("sort", "created_at")
    sort_dir = request.args.get("dir", "desc")
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"

    rows = db.query_customers(
        search=search,
        status=status_filter,
        region=region_filter,
        debt=(debt_filter == "yes"),
        sort_by=sort_by,
        sort_dir=sort_dir,
    )

    cur_month, cur_year = now_dt().month, now_dt().year
    customer_list = []
    for c in rows:
        total_debt = db.customer_unpaid_debt(c["id"])
        cur_inv = db.get_customer_invoice(c["id"], cur_month, cur_year)
        if cur_inv is None:
            current_month_paid = 0
            current_month_paid_amount = 0
        else:
            paid_c = min(float(cur_inv["paid_amount"] or 0), float(cur_inv["total_amount"] or 0))
            current_month_paid = 1 if paid_c >= float(cur_inv["total_amount"] or 0) else 0
            current_month_paid_amount = paid_c

        customer_list.append({
            "id": c["id"], "name": c["name"], "phone": c["phone"], "phone2": c["phone2"],
            "whatsapp_phone": c["whatsapp_phone"] or "",
            "address": c["address"], "region": c["region"], "username": c["username"],
            "password": c["password"] or "", "ip_address": c["ip_address"],
            "device_type": c["device_type"],
            "package_name": c["package_name"] or "", "package_price": c["package_price"] or 0,
            "renewal_date": c["renewal_date"], "subscription_status": c["subscription_status"],
            "is_active": c["is_active"], "created_at": c["created_at"],
            "subscription_date": c["subscription_date"], "notes": c["notes"],
            "total_debt": total_debt,
            "current_month_paid": current_month_paid,
            "current_month_paid_amount": current_month_paid_amount,
        })

    stats = db.customer_stats()
    regions = db.customer_regions()
    total_debt_all = db.total_unpaid_debt()

    return render_template(
        "customers.html",
        customers=customer_list, stats=stats,
        total_debt_all=total_debt_all,
        search=search, status_filter=status_filter,
        region_filter=region_filter, debt_filter=debt_filter,
        sort_by=sort_by, sort_dir=sort_dir,
        regions=regions,
    )


@app.route("/debts")
@login_required
def debts_page():
    """Customers with unpaid balances."""
    customers = db.debts_summary()
    total_debt_all = sum(c["total_debt"] for c in customers)
    return render_template("debts.html", customers=customers, total_debt_all=total_debt_all)


@app.route("/reminders")
@login_required
def reminders_page():
    """Customers who are expired or have debts — for WhatsApp deep links."""
    customers = db.customers_with_debt_or_expired()
    total_debt_all = sum(c.get("total_debt") or 0 for c in customers)
    return render_template("reminders.html", customers=customers, total_debt_all=total_debt_all)


@app.route("/payments")
@login_required
def payments_page():
    """Recent payments with filters."""
    search = request.args.get("search", "").strip()
    method_filter = request.args.get("method", "").strip()
    payments = db.list_recent_payments(search=search, method=method_filter)
    total_all = db.total_payments()
    total_today = db.total_payments_today()

    return render_template(
        "payments.html", payments=payments, search=search, method_filter=method_filter,
        total_all=total_all, total_today=total_today,
    )


@app.route("/billing")
@login_required
def billing_page():
    """Invoices for a month/year."""
    now = now_dt()
    month = request.args.get("month", now.month, type=int)
    year = request.args.get("year", now.year, type=int)
    search_bill = request.args.get("search", "").strip()

    rows = db.list_invoices(month, year, search_bill)
    inv_list = []
    for inv in rows:
        extras = db.list_invoice_extras(inv["id"])
        paid = min(float(inv["paid_amount"] or 0), float(inv["total_amount"] or 0))
        is_paid = paid >= float(inv["total_amount"] or 0)
        inv_list.append({
            "id": inv["id"], "cust_id": inv["cust_id"], "customer_name": inv["customer_name"],
            "customer_id": inv["customer_id"],
            "month": inv["month"], "year": inv["year"], "package_name": inv["package_name"],
            "package_price": inv["package_price"], "total_amount": inv["total_amount"],
            "paid_amount": paid, "is_paid": is_paid,
            "previous_debt": inv["previous_debt"],
            "extras": [{"id": e["id"], "item_name": e["item_name"], "item_price": e["item_price"]} for e in extras],
            "extras_total": sum(e["item_price"] for e in extras),
        })

    total_all = sum(i["total_amount"] for i in inv_list)
    total_paid = sum(i["paid_amount"] for i in inv_list)

    return render_template(
        "billing.html", invoices=inv_list, month=month, year=year,
        total_all=total_all, total_paid=total_paid,
        total_remaining=total_all - total_paid,
    )


@app.route("/expenses")
@login_required
def expenses_page():
    """Expenses for a month/year."""
    now = now_dt()
    month = request.args.get("month", now.month, type=int)
    year = request.args.get("year", now.year, type=int)
    search = request.args.get("search", "").strip()
    category_filter = request.args.get("category", "").strip()

    subscribers = db.list_active_customers()
    expenses = db.list_expenses(month, year, search, category_filter)
    total_expenses = db.total_expenses(month, year)
    categories = db.expense_categories(month, year)

    return render_template(
        "expenses.html", expenses=expenses, month=month, year=year,
        total_expenses=total_expenses, categories=categories,
        search=search, category_filter=category_filter,
        subscribers=subscribers,
    )


@app.route("/tickets")
@login_required
def tickets_page():
    """Maintenance tickets."""
    status_filter = request.args.get("status", "all")
    search = request.args.get("search", "").strip()
    tickets = db.list_tickets(status=status_filter, search=search)
    customers = db.list_active_customers()

    return render_template(
        "tickets.html", tickets=tickets, customers=customers,
        status_filter=status_filter, search=search,
    )


@app.route("/packages")
@login_required
def packages_page():
    """ISP packages."""
    packages = db.list_packages()
    return render_template("packages.html", packages=packages)


@app.route("/report")
@login_required
def monthly_report():
    """Monthly financial report."""
    now = now_dt()
    month = request.args.get("month", now.month, type=int)
    year = request.args.get("year", now.year, type=int)

    expected = db.sum_invoice_amounts(month, year, "total_amount")
    collected = db.sum_invoice_amounts(month, year, "paid_amount")
    expenses_total = db.total_expenses(month, year)
    remaining = expected - collected
    net_profit = collected - expenses_total
    collected_today = db.total_payments_today()
    total_debt = db.total_unpaid_debt()
    unpaid_count = db.count_unpaid_invoices(month, year)

    expense_categories = db.expense_categories(month, year)
    paid_count = db.count_paid_invoices(month, year)
    total_invoices = db.count_invoices(month, year)

    return render_template(
        "report.html", month=month, year=year,
        expected=expected, collected=collected,
        collected_today=collected_today, total_debt=total_debt,
        unpaid_count=unpaid_count,
        expenses_total=expenses_total, remaining=remaining, net_profit=net_profit,
        expense_categories=expense_categories,
        paid_count=paid_count, total_invoices=total_invoices,
    )


@app.route("/settings")
@admin_required
def settings_page():
    """Tower Settings page (admin only — agents CANNOT access)."""
    gen_info = get_gen_info()
    settings = db.get_settings()
    default_connection_type = settings.get("default_connection_type", "كيبل ضوئي")
    return render_template(
        "settings.html",
        gen_info=gen_info,
        settings=settings,
        default_connection_type=default_connection_type,
    )


@app.route("/more")
@login_required
def more_page():
    """More page (المزيد) — consolidated financial & admin tools."""
    return render_template("more.html")


@app.route("/team")
@admin_required
def team_page():
    """Team Management (الفريق) — admin only."""
    return render_template("team.html")


@app.route("/audit")
@admin_required
def audit_page():
    """Audit Log (سجل العمليات) — admin only."""
    logs = db.list_audit_logs(limit=300)
    return render_template("audit_log.html", logs=logs)


# ──────────────────────────────────────────────
#  API: CUSTOMERS
# ──────────────────────────────────────────────

@app.route("/api/customers/search")
@login_required
def api_customers_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"ok": True, "customers": []})
    customers = db.search_customers(q)
    return jsonify({
        "ok": True,
        "customers": [
            {"id": c["id"], "name": c["name"], "username": c["username"], "phone": c["phone"]}
            for c in customers
        ],
    })


def resolve_package_id(package_name):
    """Resolve a package name to its id (or None)."""
    if not package_name:
        return None
    pkg = db.get_package_by_name(package_name)
    return pkg["id"] if pkg else None


@app.route("/api/customers/add", methods=["POST"])
@admin_required
def api_customer_add():
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"ok": False, "error": "الاسم مطلوب"}), 400

    existing = db.get_customer_by_name(name)
    if existing:
        return jsonify({"ok": False, "error": "يوجد مشترك بنفس الاسم"}), 400

    try:
        int_price = int(float(data.get("package_price", 0)))
    except (ValueError, TypeError):
        int_price = 0
    months = int(data.get("duration_months", 1) or 1)
    if months <= 0:
        months = 1
    try:
        previous_debt = int(float(data.get("previous_debt", 0) or 0))
    except (ValueError, TypeError):
        previous_debt = 0
    if previous_debt < 0:
        previous_debt = 0
    # واصل = "Paid in Full" checkbox (true → debt invoice is created PAID)
    paid_in_full = bool(data.get("paid_in_full")) or previous_debt <= 0

    now = now_dt()
    from dateutil.relativedelta import relativedelta
    package_id = resolve_package_id(data.get("package_name", "").strip())

    try:
        customer_id = db.add_customer(
            full_name=name,
            phone=data.get("phone", "").strip(),
            phone2=data.get("phone2", "").strip(),
            whatsapp_phone=data.get("whatsapp_phone", "").strip(),
            address=data.get("address", "").strip(),
            region=data.get("region", "").strip(),
            package_id=package_id,
            mikrotik_username=data.get("username", "").strip(),
            mikrotik_password=data.get("password", "").strip(),
            nano_ip=data.get("ip_address", "").strip(),
            device_type=data.get("device_type", "").strip(),
            subscription_date=fmt_dt_db(now),
            renewal_date=fmt_dt_db(now + relativedelta(months=months)),
            status="active",
            previous_debt=previous_debt,
            notes=data.get("notes", "").strip(),
        )
        # Single current-month invoice: package cost + carried previous debt.
        # Utower: the debt is instant-invoiced into the same monthly bill.
        # واصل settles ONLY the previous-debt portion; the package fee remains
        # owed unless the amounts happen to be fully covered.
        package_total = int_price * months
        total_due = package_total + previous_debt
        paid_amount = previous_debt if paid_in_full else 0
        db.add_invoice(
            customer_id=customer_id, month=now.month, year=now.year,
            package_name=data.get("package_name", "").strip(),
            package_price=int_price,
            total_amount=total_due,
            paid_amount=paid_amount,
            is_paid=paid_amount >= total_due,
            previous_debt=previous_debt,
        )
        # Auto sync MikroTik: paid-in-full → enable; carrying debt → disable
        customer = db.get_customer(customer_id)
        if customer:
            sync_customer_debt(customer)
        _audit("اضافة مشترك", "customer", customer_id,
               f"تم اضافة المشترك {name} (دين سابق {previous_debt} د.ع)")
        return jsonify({"ok": True, "customer_id": customer_id})
    except Exception as e:
        log.error("Customer add failed: %s", e)
        return jsonify({"ok": False, "error": f"خطأ: {e}"}), 500


@app.route("/api/customers/edit/<int:customer_id>", methods=["POST"])
@admin_required
def api_customer_edit(customer_id):
    data = request.get_json() or {}
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404

    try:
        int_price = int(float(data.get("package_price", 0)))
    except (ValueError, TypeError):
        int_price = 0

    package_id = resolve_package_id(data.get("package_name", "").strip())

    db.update_customer(
        customer_id,
        name=data.get("name", customer["name"]).strip(),
        phone=data.get("phone", "").strip(),
        phone2=data.get("phone2", "").strip(),
        whatsapp_phone=data.get("whatsapp_phone", "").strip(),
        address=data.get("address", "").strip(),
        region=data.get("region", "").strip(),
        notes=data.get("notes", "").strip(),
        username=data.get("username", "").strip(),
        password=data.get("password", "").strip(),
        ip_address=data.get("ip_address", "").strip(),
        device_type=data.get("device_type", "").strip(),
        package_id=package_id,
        status=customer["subscription_status"],
    )
    # Keep the legacy package_price/package_name on invoices consistent on edit.
    db.update_customer_legacy_fields(
        customer_id,
        package_name=data.get("package_name", "").strip(),
        package_price=int_price,
    )
    _audit("تعديل مشترك", "customer", customer_id, f"تم تعديل بيانات المشترك {customer['name']}")
    return jsonify({"ok": True})


@app.route("/api/customers/toggle/<int:customer_id>", methods=["POST"])
@admin_required
def api_customer_toggle(customer_id):
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404
    db.toggle_customer(customer_id)
    _audit("تبديل تفعيل مشترك", "customer", customer_id, f"تم تبديل حالة المشترك {customer['name']}")
    return jsonify({"ok": True})


@app.route("/api/customers/status/<int:customer_id>", methods=["POST"])
@admin_required
def api_customer_status(customer_id):
    data = request.get_json() or {}
    new_status = data.get("status", "active")
    if new_status not in ("active", "expired", "suspended"):
        return jsonify({"ok": False, "error": "حالة غير صالحة"}), 400
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404
    db.set_customer_status(customer_id, new_status)
    _audit("تغيير حالة مشترك", "customer", customer_id,
           f"تم تغيير حالة المشترك {customer['name']} إلى {new_status}")
    return jsonify({"ok": True})


@app.route("/api/customers/renew/<int:customer_id>", methods=["POST"])
@login_required
def api_customer_renew(customer_id):
    """Renew subscription by N months, appending to current expiry if in future."""
    data = request.get_json() or {}
    try:
        months = int(data.get("months", 1))
        if months <= 0:
            months = 1
    except (ValueError, TypeError):
        months = 1

    from dateutil.relativedelta import relativedelta
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404

    now = now_dt()
    base = None
    if customer["renewal_date"]:
        for fmt in (DATETIME_DB, DATETIME_DB_S):
            try:
                base = datetime.strptime(str(customer["renewal_date"]), fmt)
                break
            except ValueError:
                continue
    if base is None or base < now:
        base = now

    new_renewal = base + relativedelta(months=months)
    try:
        db.update_customer(
            customer_id,
            renewal_date=fmt_dt_db(new_renewal),
            subscription_status="active",
            status="active",
        )
        # Utower invoicing engine: payable = package x months + carried debt
        renew_total = (customer["package_price"] or 0) * months
        carried_debt = db.customer_unpaid_debt(
            customer_id, up_to_month=now.month, up_to_year=now.year
        )
        total_due = renew_total + carried_debt
        db.add_invoice(
            customer_id=customer_id, month=now.month, year=now.year,
            package_name=customer["package_name"] or "",
            package_price=customer["package_price"] or 0,
            total_amount=total_due, paid_amount=0, is_paid=False,
            previous_debt=carried_debt,
        )
        # Auto sync: any unpaid balance (incl. carried debt) disables subscriber
        refreshed = db.get_customer(customer_id)
        if refreshed:
            sync_customer_debt(refreshed)
        _audit("تجديد اشتراك", "customer", customer_id,
               f"تم تجديد اشتراك {customer['name']} لمدة {months} شهر "
               f"(اجمالي {total_due} د.ع)")
        return jsonify({
            "ok": True,
            "renewal_date": fmt_dt_db(new_renewal),
            "invoices_generated": 1,
            "total_amount": total_due,
            "carried_debt": carried_debt,
        })
    except Exception as e:
        log.error("Renew failed: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/customers/delete/<int:customer_id>", methods=["POST"])
@admin_required
def api_customer_delete(customer_id):
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404
    try:
        db.delete_customer(customer_id)
        _audit("حذف مشترك", "customer", customer_id, f"تم حذف المشترك {customer['name']}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": f"فشل الحذف: {e}"}), 500


@app.route("/api/customers/history/<int:customer_id>")
@login_required
def api_customer_history(customer_id):
    if not db.get_customer(customer_id):
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404

    invoices = db.list_customer_invoices(customer_id)
    payments = db.list_customer_payments(customer_id)

    inv_list = []
    for inv in invoices:
        extras = db.list_invoice_extras(inv["id"])
        paid = min(float(inv["paid_amount"] or 0), float(inv["total_amount"] or 0))
        is_paid = paid >= float(inv["total_amount"] or 0)
        inv_list.append({
            "id": inv["id"], "month": inv["month"], "year": inv["year"],
            "package_name": inv["package_name"] or "", "package_price": inv["package_price"] or 0,
            "total_amount": inv["total_amount"], "paid_amount": paid,
            "is_paid": is_paid,
            "extras": [
                {"id": e["id"], "item_name": e["item_name"], "item_price": e["item_price"]}
                for e in extras
            ],
        })

    pay_list = [{
        "id": p["id"], "amount": p["amount"], "payment_date": p["payment_date"],
        "payment_method": p["payment_method"], "notes": p["notes"],
    } for p in payments]

    return jsonify({"ok": True, "invoices": inv_list, "payments": pay_list})


# ──────────────────────────────────────────────
#  API: PAYMENTS
# ──────────────────────────────────────────────

@app.route("/api/payment/quick-pay/<int:customer_id>", methods=["POST"])
@login_required
def api_quick_pay(customer_id):
    data = request.get_json() or {}
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404

    now = now_dt()
    month, year = now.month, now.year

    invoice = db.get_customer_invoice(customer_id, month, year)
    if invoice is None:
        total = customer["package_price"] or 0
        invoice_id = db.add_invoice(
            customer_id=customer_id, month=month, year=year,
            package_name=customer["package_name"] or "",
            package_price=customer["package_price"] or 0,
            total_amount=total, paid_amount=0, is_paid=False,
        )
        invoice = db.get_invoice(invoice_id)

    remaining = (invoice["total_amount"] or 0) - (invoice["paid_amount"] or 0)
    if remaining <= 0:
        return jsonify({"ok": False, "error": "الفاتورة مدفوعة بالكامل"}), 400

    try:
        pay_amount = float(data.get("amount")) if data.get("amount") else remaining
    except (ValueError, TypeError):
        pay_amount = remaining
    if pay_amount > remaining:
        pay_amount = remaining
    if pay_amount <= 0:
        return jsonify({"ok": False, "error": "المبلغ غير صالح"}), 400

    new_paid = (invoice["paid_amount"] or 0) + pay_amount
    is_paid = new_paid >= (invoice["total_amount"] or 0)
    db.update_invoice(invoice["id"], paid_amount=new_paid, is_paid=is_paid)

    pay_date_str = data.get("payment_date", now.strftime("%Y-%m-%d"))
    try:
        pay_date = datetime.strptime(pay_date_str, "%Y-%m-%d").date().isoformat()
    except ValueError:
        pay_date = date.today().isoformat()

    db.add_payment(
        invoice_id=invoice["id"], customer_id=customer_id,
        amount=pay_amount, payment_date=pay_date,
        payment_method=data.get("payment_method", "نقدي"),
        notes=data.get("notes", "").strip(),
    )
    _audit("دفع", "payment", invoice["id"],
           f"تم استلام {pay_amount} د.ع من {customer['name']}")
    # Utower: cleared ALL debt → enable router; still owes → disable router
    refreshed = db.get_customer(customer_id)
    if refreshed:
        sync_customer_debt(refreshed)
    return jsonify({
        "ok": True, "invoice_id": invoice["id"], "amount": pay_amount,
        "remaining": (invoice["total_amount"] or 0) - new_paid,
    })


@app.route("/api/payment/make", methods=["POST"])
@login_required
def api_make_payment():
    data = request.get_json() or {}
    invoice_id = data.get("invoice_id")

    try:
        amount = float(data.get("amount", 0))
    except (ValueError, TypeError):
        amount = 0
    if not invoice_id or amount <= 0:
        return jsonify({"ok": False, "error": "البيانات غير صالحة"}), 400

    invoice = db.get_invoice(invoice_id)
    if not invoice:
        return jsonify({"ok": False, "error": "الفاتورة غير موجودة"}), 404

    remaining = (invoice["total_amount"] or 0) - (invoice["paid_amount"] or 0)
    if amount > remaining:
        amount = remaining

    new_paid = (invoice["paid_amount"] or 0) + amount
    is_paid = new_paid >= (invoice["total_amount"] or 0)
    db.update_invoice(invoice_id, paid_amount=new_paid, is_paid=is_paid)

    pay_date_str = data.get("payment_date", now_dt().strftime("%Y-%m-%d"))
    try:
        pay_date = datetime.strptime(pay_date_str, "%Y-%m-%d").date().isoformat()
    except ValueError:
        pay_date = date.today().isoformat()

    db.add_payment(
        invoice_id=invoice_id, customer_id=invoice["customer_id"],
        amount=amount, payment_date=pay_date,
        payment_method=data.get("payment_method", "نقدي"),
        notes=data.get("notes", "").strip(),
    )
    _audit("تسديد فاتورة", "payment", invoice_id,
           f"تم تسديد {amount} د.ع على فاتورة رقم {invoice_id}")
    # Utower: after payment, re-sync router status from the customer's debt
    refreshed = db.get_customer(invoice["customer_id"])
    if refreshed:
        sync_customer_debt(refreshed)
    return jsonify({
        "ok": True, "paid_amount": new_paid,
        "remaining": (invoice["total_amount"] or 0) - new_paid,
    })


@app.route("/api/payment/edit/<int:payment_id>", methods=["POST"])
@login_required
def api_payment_edit(payment_id):
    data = request.get_json() or {}
    payment = db.get_payment(payment_id)
    if not payment:
        return jsonify({"ok": False, "error": "الدفعة غير موجودة"}), 404

    try:
        new_amount = float(data.get("amount", 0))
    except (ValueError, TypeError):
        new_amount = 0
    if new_amount <= 0:
        return jsonify({"ok": False, "error": "المبلغ غير صالح"}), 400

    invoice = db.get_invoice(payment["invoice_id"])
    if not invoice:
        return jsonify({"ok": False, "error": "الفاتورة غير موجودة"}), 404

    old_amount = payment["amount"]
    new_paid = (invoice["paid_amount"] or 0) - old_amount + new_amount
    if new_paid > invoice["total_amount"]:
        new_paid = invoice["total_amount"]
    if new_paid < 0:
        new_paid = 0
    is_paid = new_paid >= (invoice["total_amount"] or 0)
    db.update_invoice(invoice["id"], paid_amount=new_paid, is_paid=is_paid)

    pay_date = payment["payment_date"]
    try:
        pay_date = datetime.strptime(data.get("payment_date", ""), "%Y-%m-%d").date().isoformat()
    except ValueError:
        pass

    db.update_payment(
        payment_id,
        amount=new_amount,
        payment_date=pay_date,
        payment_method=data.get("payment_method", payment["payment_method"]),
        notes=data.get("notes", "").strip(),
    )
    return jsonify({"ok": True, "remaining": (invoice["total_amount"] or 0) - new_paid})


@app.route("/api/payment/delete/<int:payment_id>", methods=["POST"])
@login_required
def api_payment_delete(payment_id):
    payment = db.get_payment(payment_id)
    if not payment:
        return jsonify({"ok": False, "error": "الدفعة غير موجودة"}), 404

    invoice = db.get_invoice(payment["invoice_id"])
    if invoice:
        new_paid = (invoice["paid_amount"] or 0) - payment["amount"]
        if new_paid < 0:
            new_paid = 0
        is_paid = new_paid >= (invoice["total_amount"] or 0)
        db.update_invoice(invoice["id"], paid_amount=new_paid, is_paid=is_paid)

    db.delete_payment(payment_id)
    return jsonify({"ok": True})


@app.route("/api/payment/current-invoice/<int:customer_id>")
@login_required
def api_current_invoice(customer_id):
    now = now_dt()
    invoice = db.get_customer_invoice(customer_id, now.month, now.year)
    if invoice:
        return jsonify({
            "ok": True,
            "invoice": {
                "id": invoice["id"],
                "total_amount": invoice["total_amount"],
                "paid_amount": invoice["paid_amount"],
            },
            "remaining": (invoice["total_amount"] or 0) - (invoice["paid_amount"] or 0),
            "paid": invoice["paid_amount"], "total": invoice["total_amount"],
        })
    return jsonify({"ok": True, "invoice": None, "remaining": 0, "paid": 0, "total": 0})


# ──────────────────────────────────────────────
#  API: EXPENSES
# ──────────────────────────────────────────────

@app.route("/api/expenses/add", methods=["POST"])
@login_required
def api_expense_add():
    data = request.get_json() or {}
    try:
        amount = float(data.get("amount", 0))
    except (ValueError, TypeError):
        amount = 0
    if amount <= 0:
        return jsonify({"ok": False, "error": "المبلغ يجب أن يكون أكبر من صفر"}), 400
    category = data.get("category", "أخرى").strip()
    if not category:
        return jsonify({"ok": False, "error": "التصنيف مطلوب"}), 400

    try:
        expense_date = datetime.strptime(data.get("expense_date", ""), "%Y-%m-%d").date().isoformat()
    except ValueError:
        expense_date = date.today().isoformat()

    subscriber_id = data.get("subscriber_id")
    if subscriber_id in ("", None, "null"):
        subscriber_id = None

    try:
        db.add_expense(
            expense_date=expense_date, category=category, amount=amount,
            description=data.get("description", "").strip(),
            recipient_name=data.get("recipient_name", "").strip(),
            subscriber_id=int(subscriber_id) if subscriber_id else None,
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/expenses/edit/<int:expense_id>", methods=["POST"])
@login_required
def api_expense_edit(expense_id):
    data = request.get_json() or {}
    ex = db.get_expense(expense_id)
    if not ex:
        return jsonify({"ok": False, "error": "المصروف غير موجود"}), 404

    try:
        amount = float(data.get("amount", 0))
    except (ValueError, TypeError):
        amount = 0
    if amount <= 0:
        return jsonify({"ok": False, "error": "المبلغ يجب أن يكون أكبر من صفر"}), 400

    expense_date = ex["expense_date"]
    try:
        expense_date = datetime.strptime(data.get("expense_date", ""), "%Y-%m-%d").date().isoformat()
    except ValueError:
        pass

    sub_id = data.get("subscriber_id")
    subscriber_id = int(sub_id) if sub_id not in ("", None, "null") else None

    db.update_expense(
        expense_id,
        amount=amount,
        category=data.get("category", ex["category"]).strip(),
        description=data.get("description", "").strip(),
        recipient_name=data.get("recipient_name", "").strip(),
        expense_date=expense_date,
        subscriber_id=subscriber_id,
    )
    return jsonify({"ok": True})


@app.route("/api/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
def api_expense_delete(expense_id):
    if not db.get_expense(expense_id):
        return jsonify({"ok": False, "error": "المصروف غير موجود"}), 404
    db.delete_expense(expense_id)
    return jsonify({"ok": True})


# ──────────────────────────────────────────────
#  API: PACKAGES
# ──────────────────────────────────────────────

@app.route("/api/packages/list")
@login_required
def api_packages_list():
    packages = db.list_packages()
    return jsonify({
        "ok": True,
        "packages": [
            {"id": p["id"], "name": p["name"], "price": p["price"], "speed": p["speed"] or ""}
            for p in packages
        ],
    })


@app.route("/api/packages/add", methods=["POST"])
@admin_required
def api_packages_add():
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    try:
        price = int(float(data.get("price", 0)))
    except (ValueError, TypeError):
        price = 0
    if not name or price <= 0:
        return jsonify({"ok": False, "error": "البيانات غير صالحة"}), 400

    if db.get_package_by_name(name):
        return jsonify({"ok": False, "error": "يوجد باقة بنفس الاسم"}), 400

    db.add_package(name=name, price=price, speed=data.get("speed", "").strip())
    _audit("اضافة باقة", "package", None, f"تم اضافة الباقة {name} بسعر {price} د.ع")
    return jsonify({"ok": True})


@app.route("/api/packages/edit/<int:package_id>", methods=["POST"])
@admin_required
def api_packages_edit(package_id):
    data = request.get_json() or {}
    pkg = db.get_package(package_id)
    if not pkg:
        return jsonify({"ok": False, "error": "الباقة غير موجودة"}), 404

    try:
        price = int(float(data.get("price", 0)))
    except (ValueError, TypeError):
        price = 0

    new_name = data.get("name", pkg["name"]).strip()
    if new_name != pkg["name"] and db.get_package_by_name(new_name):
        return jsonify({"ok": False, "error": "يوجد باقة بنفس الاسم"}), 400

    db.update_package(
        package_id,
        name=new_name,
        price=price,
        speed=data.get("speed", pkg["speed"] or "").strip(),
    )
    _audit("تعديل باقة", "package", package_id, f"تم تعديل الباقة {new_name}")
    return jsonify({"ok": True})


@app.route("/api/packages/delete/<int:package_id>", methods=["POST"])
@admin_required
def api_packages_delete(package_id):
    pkg = db.get_package(package_id)
    if not pkg:
        return jsonify({"ok": False, "error": "الباقة غير موجودة"}), 404
    db.delete_package(package_id)
    _audit("حذف باقة", "package", package_id, f"تم حذف الباقة {pkg['name']}")
    return jsonify({"ok": True})


# ──────────────────────────────────────────────
#  API: INVOICING & TICKETS
# ──────────────────────────────────────────────

@app.route("/api/billing/generate", methods=["POST"])
@login_required
def api_billing_generate():
    data = request.get_json() or {}
    now = now_dt()
    month = int(data.get("month", now.month))
    year = int(data.get("year", now.year))

    customers = db.list_active_customers()
    generated = skipped = 0
    for c in customers:
        if db.get_customer_invoice(c["id"], month, year):
            skipped += 1
            continue
        db.add_invoice(
            customer_id=c["id"], month=month, year=year,
            package_name=c["package_name"] or "",
            package_price=c["package_price"] or 0,
            total_amount=c["package_price"] or 0,
            paid_amount=0, is_paid=False,
        )
        generated += 1
    return jsonify({"ok": True, "generated": generated, "skipped": skipped})


@app.route("/api/invoice/extras/add", methods=["POST"])
@login_required
def api_invoice_extra_add():
    data = request.get_json() or {}
    invoice_id = data.get("invoice_id")
    item_name = data.get("item_name", "").strip()
    try:
        item_price = int(float(data.get("item_price", 0)))
    except (ValueError, TypeError):
        item_price = 0
    if not invoice_id or not item_name or item_price <= 0:
        return jsonify({"ok": False, "error": "البيانات غير صالحة"}), 400

    inv = db.get_invoice(invoice_id)
    if not inv:
        return jsonify({"ok": False, "error": "الفاتورة غير موجودة"}), 404

    db.add_invoice_extra(invoice_id, item_name, item_price)
    extras_total = db.extras_total(invoice_id)
    new_total = (inv["package_price"] or 0) + extras_total
    paid = inv["paid_amount"] or 0
    if paid > new_total:
        paid = new_total
    is_paid = paid >= new_total
    db.update_invoice(invoice_id, total_amount=new_total, paid_amount=paid, is_paid=is_paid)
    return jsonify({"ok": True, "new_total": new_total})


@app.route("/api/invoice/extras/delete/<int:extra_id>", methods=["POST"])
@login_required
def api_invoice_extra_delete(extra_id):
    extra = db.get_invoice_extra(extra_id)
    if not extra:
        return jsonify({"ok": False, "error": "العنصر غير موجود"}), 404

    invoice_id = extra["invoice_id"]
    inv = db.get_invoice(invoice_id)
    db.delete_invoice_extra(extra_id)

    if inv:
        extras_total = db.extras_total(invoice_id)
        new_total = (inv["package_price"] or 0) + extras_total
        paid = inv["paid_amount"] or 0
        if paid > new_total:
            paid = new_total
        is_paid = paid >= new_total
        db.update_invoice(invoice_id, total_amount=new_total, paid_amount=paid, is_paid=is_paid)
    return jsonify({"ok": True})


@app.route("/api/tickets/add", methods=["POST"])
@login_required
def api_ticket_add():
    data = request.get_json() or {}
    customer_id = data.get("customer_id")
    issue = data.get("issue_description", "").strip()
    if not customer_id:
        return jsonify({"ok": False, "error": "يرجى اختيار المشترك"}), 400
    if not issue:
        return jsonify({"ok": False, "error": "يرجى كتابة وصف المشكلة"}), 400

    if not db.get_customer(customer_id):
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404

    db.add_ticket(customer_id, issue)
    return jsonify({"ok": True})


@app.route("/api/tickets/resolve/<int:ticket_id>", methods=["POST"])
@login_required
def api_ticket_resolve(ticket_id):
    if not db.get_ticket(ticket_id):
        return jsonify({"ok": False, "error": "طلب الصيانة غير موجود"}), 404
    new_status = db.toggle_ticket(ticket_id)
    return jsonify({"ok": True, "new_status": new_status})


@app.route("/api/tickets/delete/<int:ticket_id>", methods=["POST"])
@login_required
def api_ticket_delete(ticket_id):
    if not db.get_ticket(ticket_id):
        return jsonify({"ok": False, "error": "طلب الصيانة غير موجود"}), 404
    db.delete_ticket(ticket_id)
    return jsonify({"ok": True})


@app.route("/api/billing/rollover", methods=["POST"])
@login_required
def api_debt_rollover():
    """Roll unpaid previous debt into the target month's invoices."""
    data = request.get_json() or {}
    now = now_dt()
    target_month = int(data.get("month", now.month))
    target_year = int(data.get("year", now.year))

    from sqlalchemy import or_, and_  # noqa: F401  (kept for parity; not used)
    inv_count = db.count_invoices(target_month, target_year)
    if inv_count == 0:
        return jsonify({
            "ok": False,
            "error": "لا توجد فواتير للشهر المحدد.",
            "code": "NO_INVOICES",
        }), 400

    customers = db.list_active_customers()
    rolled = 0
    for c in customers:
        prev_debt = db.customer_unpaid_debt(
            c["id"], up_to_month=target_month, up_to_year=target_year
        )
        if prev_debt <= 0:
            continue

        inv = db.get_customer_invoice(c["id"], target_month, target_year)
        if inv:
            db.update_invoice(inv["id"], previous_debt=prev_debt)
        else:
            db.add_invoice(
                customer_id=c["id"], month=target_month, year=target_year,
                package_name=c["package_name"] or "",
                package_price=c["package_price"] or 0,
                total_amount=c["package_price"] or 0,
                paid_amount=0, is_paid=False,
                previous_debt=prev_debt,
            )
        rolled += 1

    return jsonify({"ok": True, "rolled": rolled})


# ──────────────────────────────────────────────
#  API: SETTINGS
# ──────────────────────────────────────────────

@app.route("/api/settings/numeral-style", methods=["POST"])
@admin_required
def api_numeral_style():
    data = request.get_json() or {}
    style = data.get("numeral_style", "AR")
    if style not in ("AR", "EN"):
        return jsonify({"ok": False, "error": "قيمة غير صالحة"}), 400
    db.set_setting("numeral_style", style)
    _audit("تغيير نمط الأرقام", "settings", None, f"تم تغيير نمط الأرقام إلى {style}")
    return jsonify({"ok": True})


@app.route("/api/settings/default-connection", methods=["POST"])
@admin_required
def api_settings_default_connection():
    data = request.get_json() or {}
    conn_type = data.get("default_connection_type", "كيبل ضوئي")
    if conn_type not in ("كيبل ضوئي", "نانو"):
        return jsonify({"ok": False, "error": "نوع ربط غير صالح"}), 400
    db.set_setting("default_connection_type", conn_type)
    _audit("تغيير نوع الربط", "settings", None, f"تم تغيير نوع الربط إلى {conn_type}")
    return jsonify({"ok": True})


@app.route("/api/generator/update", methods=["POST"])
@admin_required
def api_isp_info_update():
    data = request.get_json() or {}
    db.update_generator_info(
        owner_name=data.get("owner_name", "").strip() or None,
        phone=data.get("phone", "").strip() or None,
        address=data.get("address", "").strip() or None,
        footer_note=data.get("footer_note", "").strip() or None,
    )
    _audit("تعديل بيانات الشركة", "settings", None, "تم تعديل بيانات الشركة في الإعدادات")
    return jsonify({"ok": True})


# ──────────────────────────────────────────────
#  API: NETWORK TOOLS (Phase 13)
# ──────────────────────────────────────────────

@app.route("/network")
@admin_required
def network_page():
    """Network Tools & Tower Management page (admin only)."""
    links = db.list_network_links()
    settings = db.get_settings()
    return render_template("network.html", links=links, settings=settings)


@app.route("/api/settings/tower-connection", methods=["POST"])
@admin_required
def api_tower_connection_save():
    """Save MikroTik + OLT/SNMP connection settings into the settings table."""
    data = request.get_json() or {}
    mikrotik_host = data.get("mikrotik_host", "").strip()
    if mikrotik_host and not validate_host(mikrotik_host):
        return jsonify({"ok": False, "error": "عنوان MikroTik غير صالح"}), 400

    olt_ip = data.get("olt_ip", "").strip()
    if olt_ip and not validate_host(olt_ip):
        return jsonify({"ok": False, "error": "عنوان OLT غير صالح"}), 400

    try:
        port = int(data.get("mikrotik_port", 8728) or 8728)
        if not 1 <= port <= 65535:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "منفذ MikroTik غير صالح"}), 400

    try:
        snmp_port = int(data.get("snmp_port", 161) or 161)
        if not 1 <= snmp_port <= 65535:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "منفذ SNMP غير صالح"}), 400

    try:
        snmp_timeout = float(data.get("snmp_timeout", 2.0) or 2.0)
        if not 0.5 <= snmp_timeout <= 30:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "مهلة SNMP غير صالحة"}), 400

    try:
        snmp_retries = int(data.get("snmp_retries", 1) or 1)
        if not 0 <= snmp_retries <= 10:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "عدد محاولات SNMP غير صالح"}), 400

    keys = {
        "mikrotik_host": mikrotik_host,
        "mikrotik_user": data.get("mikrotik_user", "").strip(),
        "mikrotik_password": data.get("mikrotik_password", ""),
        "mikrotik_port": str(port),
        "mikrotik_ssl": "1" if data.get("mikrotik_ssl") else "0",
        "olt_ip": olt_ip,
        "olt_snmp_community": data.get("olt_snmp_community", "").strip(),
        "snmp_community": data.get("snmp_community", "public").strip(),
        "snmp_port": str(snmp_port),
        "snmp_timeout": str(snmp_timeout),
        "snmp_retries": str(snmp_retries),
        "oid_onu_rx": data.get("oid_onu_rx", "").strip(),
        "oid_onu_tx": data.get("oid_onu_tx", "").strip(),
        "oid_ubnt_signal": data.get("oid_ubnt_signal", "").strip(),
        "oid_ubnt_ccq": data.get("oid_ubnt_ccq", "").strip(),
        "oid_mikrotik_signal": data.get("oid_mikrotik_signal", "").strip(),
    }
    for key, value in keys.items():
        db.set_setting(key, value)

    _audit("حفظ اعدادات اللوحة", "settings", None,
           f"تم حفظ اعدادات ربط اللوحة (MikroTik: {mikrotik_host})")
    return jsonify({"ok": True, "message": "تم حفظ اعدادات اللوحة بنجاح"})


@app.route("/api/settings/tower-test", methods=["POST"])
@admin_required
def api_tower_connection_test():
    """Attempt a live connection to the saved MikroTik router."""
    from mikrotik_api.config import load_mikrotik_config
    cfg = load_mikrotik_config()
    if not cfg.get("host"):
        return jsonify({"ok": False, "error": "لم يتم حفظ عنوان MikroTik بعد"}), 400
    try:
        manager = MikroTikManager(
            host=cfg["host"],
            username=cfg["username"],
            password=cfg["password"],
            port=cfg["port"],
        )
        with manager:
            manager.get_ppp_secrets()
        return jsonify({"ok": True, "message": "تم الاتصال بالراوتر بنجاح"})
    except Exception as e:
        log.error("Tower connection test failed: %s", e)
        return jsonify({"ok": False, "error": "فشل الاتصال بالراوتر: " + str(e)}), 400


@app.route("/api/network/ping", methods=["POST"])
@admin_required
def api_network_ping():
    """Run a live ICMP ping against a target IP/host."""
    data = request.get_json() or {}
    host = data.get("host", "").strip()
    try:
        count = int(data.get("count", 4))
    except (ValueError, TypeError):
        count = 4
    log.info("PING API host=%s count=%s", host, count)
    result = ping_host(host, count=count)
    return jsonify(result)


@app.route("/signal-board")
@login_required
def signal_board_page():
    """All-subscribers signal board (local tower PC + remote views)."""
    return render_template("signal_board.html")


@app.route("/api/signal-board-data")
@login_required
def api_signal_board_data():
    """Return cached signals joined with customer names for the board."""
    cache_rows = db.list_signal_cache()
    by_ip = {r["ip"]: dict(r) for r in cache_rows}
    items = []
    last_update = ""
    for c in db.list_active_customers():
        ip = (c.get("ip_address") or "").strip()
        row = by_ip.get(ip)
        if not row:
            continue
        dt = (c.get("device_type") or "").strip()
        if dt == "كيبل ضوئي":
            row["type"] = "optical"
        else:
            row["type"] = "wireless"
        row["name"] = c["name"]
        if row.get("last_updated") and row["last_updated"] > last_update:
            last_update = row["last_updated"]
        items.append(row)
    # Also include cached entries whose customer is no longer active (still useful).
    known = {i["ip"] for i in items}
    for ip, row in by_ip.items():
        if ip in known:
            continue
        row["type"] = "wireless"
        row["name"] = ip
        items.append(row)
    return jsonify({"ok": True, "items": items, "last_update": last_update})


@app.route("/api/sync/pull", methods=["POST"])
@admin_required
def api_sync_pull():
    """Pull MikroTik router data into the local DB, then push to cloud.

    Hybrid execution (Local Laptop/PC): this endpoint:
      1. Connects to the MikroTik router over the LAN,
      2. Auto-maps packages + upserts subscribers (sync_pull_router),
      3. Pushes the resulting snapshot to Render via /api/sync/push
         (sync_push_cloud) when RELAY_URL is configured.

    Returns:
        JSON with {pull, push} summaries.
    """
    from billing_system.mikrotik_sync import sync_pull_router, sync_push_cloud

    try:
        pull = sync_pull_router(dry_run=False)
        push = {}
        if pull.get("secrets", 0) > 0:
            push = sync_push_cloud()
        return jsonify({"ok": True, "pull": pull,
                        "push": push,
                        "message": "تم سحب بيانات الراوتر ومزامنة السحابة"})
    except Exception as e:
        log.error("Hybrid sync failed: %s", e)
        return jsonify({"ok": False, "error": f"فشل المزامنة: {e}"}), 500


@app.route("/api/sync/push", methods=["POST"])
def api_sync_push():
    """Accept a full DB snapshot pushed from a local laptop/PC (hybrid upload).

    Guarded by the same Bearer token as the agent signal relay. Upserts
    packages + customers into the cloud DB so remote/mobile views see the
    local tower snapshot instantly.
    """
    from config import AGENT_TOKEN

    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    if not token or token != AGENT_TOKEN:
        return jsonify({"ok": False, "error": "غير مصرح"}), 401

    data = request.get_json(silent=True) or {}
    customers = data.get("customers") or []
    packages = data.get("packages") or []
    upserted = 0

    # Upsert packages first so customer FK references resolve.
    for p in packages:
        name = (p.get("name") or "").strip()
        if not name:
            continue
        existing = db.get_package_by_name(name)
        if existing:
            db.update_package(existing["id"], price=p.get("price", 0), speed=p.get("speed", ""))
        else:
            db.add_package(name=name, price=p.get("price", 0), speed=p.get("speed", ""))

    # Upsert customers by mikrotik_username (single subscriber identity).
    for c in customers:
        username = (c.get("username") or "").strip()
        if not username:
            continue
        pkg = None
        pkg_name = (c.get("package_name") or "").strip()
        if pkg_name:
            p = db.get_package_by_name(pkg_name)
            pkg = p["id"] if p else None
        rows = db._fetchall("SELECT id FROM customers WHERE mikrotik_username = ?", (username,))
        if rows:
            db.update_customer(
                rows[0]["id"],
                name=c.get("name", ""),
                phone=c.get("phone", ""),
                phone2=c.get("phone2", ""),
                whatsapp_phone=c.get("whatsapp_phone", ""),
                address=c.get("address", ""),
                region=c.get("region", ""),
                package_id=pkg,
                username=username,
                ip_address=c.get("ip_address", ""),
                device_type=c.get("device_type", ""),
                status=c.get("subscription_status", "active"),
                subscription_status=c.get("subscription_status", "active"),
                renewal_date=c.get("renewal_date", ""),
                notes=c.get("notes", ""),
            )
        else:
            db.add_customer(
                full_name=c.get("name") or username,
                phone=c.get("phone", ""),
                phone2=c.get("phone2", ""),
                whatsapp_phone=c.get("whatsapp_phone", ""),
                address=c.get("address", ""),
                region=c.get("region", ""),
                package_id=pkg,
                mikrotik_username=username,
                mikrotik_password=c.get("password", ""),
                nano_ip=c.get("ip_address", ""),
                device_type=c.get("device_type", ""),
                subscription_date=c.get("subscription_date", ""),
                renewal_date=c.get("renewal_date", ""),
                status=c.get("subscription_status", "active"),
            )
        upserted += 1

    _audit("مزامنة سحابية", "sync", None,
           f"تم استلام {len(packages)} باقة و {len(customers)} مشترك من الجهاز المحلي")
    return jsonify({"ok": True, "customers_upserted": upserted,
                    "packages_received": len(packages)})


@app.route("/api/agent/signal", methods=["POST"])
def api_agent_signal():
    """Accept ONE batched signal snapshot from the tower-LAN scanner.

    Guardrails:
      - requires `Authorization: Bearer <AGENT_TOKEN>` (missing/invalid → 401).
      - accepts a single JSON batch (never per-subscriber requests).
    """
    from config import AGENT_TOKEN

    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    if not token or token != AGENT_TOKEN:
        return jsonify({"ok": False, "error": "غير مصرح"}), 401

    data = request.get_json(silent=True) or {}
    batch = data.get("batch") or []
    if not isinstance(batch, list) or not batch:
        return jsonify({"ok": False, "error": "batch فارغ"}), 400

    db.upsert_signal_batch(batch)
    return jsonify({"ok": True, "stored": len(batch)})


@app.route("/api/network/signal", methods=["POST"])
@admin_required
def api_network_signal():
    """Fetch the live signal of any Nano/optical device via SNMP.

    Prefers the agent-cached reading (from the tower-LAN scanner) when one
    exists, so remote/mobile views show real CCQ/dBm even for private IPs.
    """
    data = request.get_json() or {}
    ip = data.get("ip", "").strip()
    device_type = data.get("device_type", "auto").strip().lower()
    cached = db.get_cached_signal(ip)
    if cached and cached["last_updated"]:
        return jsonify({
            "ok": cached["status"] == "good",
            "ip": ip, "status": cached["status"],
            "signal_dbm": cached["signal_dbm"] or None,
            "ccq": cached["ccq"] or None,
            "rx_dbm": cached["rx_dbm"] or None,
            "tx_dbm": cached["tx_dbm"] or None,
            "last_updated": cached["last_updated"],
            "from_cache": True,
        })
    if not ip:
        return jsonify({"ok": False, "error": "عنوان IP مطلوب"}), 400
    if not validate_host(ip):
        return jsonify({"ok": False, "error": "عنوان IP غير صالح"}), 400
    if device_type not in ("auto", "optical", "wireless"):
        device_type = "auto"
    community = data.get("community", "public").strip() or "public"

    # Always attempt a live SNMP probe — no IP pre-block. The configured
    # timeout handles unreachable hosts naturally (offline/timeout result).
    try:
        monitor = SignalMonitor()
        signal = monitor.get_client_signal(ip, device_type, community=community)
        signal["ok"] = signal.get("status") == "good"
        return jsonify(signal)
    except Exception as e:
        log.error("SNMP signal fetch failed for %s: %s", ip, e)
        return jsonify({
            "ok": False, "ip": ip, "status": "error",
            "error": f"فشل قراءة الإشارة: {e}",
        }), 500


@app.route("/api/customers/signal/<int:customer_id>")
@admin_required
def api_customer_signal(customer_id):
    """Fetch live signal for a specific customer from their Nano IP."""
    customer = db.get_customer(customer_id)
    if not customer:
        return jsonify({"ok": False, "error": "المشترك غير موجود"}), 404
    ip = (customer.get("ip_address") or "").strip()
    if not ip:
        return jsonify({"ok": False, "error": "لا يوجد عنوان IP لهذا المشترك"}), 400
    device_type = (customer.get("device_type") or "").strip()
    if device_type == "نانو":
        device_type = "wireless"
    elif device_type == "كيبل ضوئي":
        device_type = "optical"
    else:
        device_type = "auto"

    # Prefer the agent-cached reading when present.
    cached = db.get_cached_signal(ip)
    if cached and cached["last_updated"]:
        return jsonify({
            "ok": cached["status"] == "good",
            "ip": ip, "status": cached["status"],
            "signal_dbm": cached["signal_dbm"] or None,
            "ccq": cached["ccq"] or None,
            "rx_dbm": cached["rx_dbm"] or None,
            "tx_dbm": cached["tx_dbm"] or None,
            "last_updated": cached["last_updated"],
            "from_cache": True,
            "customer_name": customer["name"],
        })

    # Always attempt a live SNMP probe — no IP pre-block. The configured
    # timeout handles unreachable hosts naturally (offline/timeout result).
    try:
        monitor = SignalMonitor()
        signal = monitor.get_client_signal(ip, device_type)
        signal["ok"] = signal.get("status") == "good"
        signal["customer_name"] = customer["name"]
        return jsonify(signal)
    except Exception as e:
        log.error("Customer signal fetch failed for %s: %s", ip, e)
        return jsonify({
            "ok": False, "ip": ip, "status": "error",
            "error": f"فشل قراءة الإشارة: {e}",
        }), 500


# ── Network Links CRUD (sectors & links) ────────────────────

@app.route("/api/network/links")
@admin_required
def api_network_links_list():
    """List all network links (sectors/links)."""
    links = db.list_network_links()
    return jsonify({
        "ok": True,
        "links": [
            {
                "id": l["id"], "name": l["name"], "ip": l["ip"] or "",
                "link_type": l["link_type"], "location": l["location"] or "",
                "notes": l["notes"] or "", "created_at": l["created_at"],
                "username": l["username"] or "", "password": l["password"] or "",
                "community": l["community"] or "public",
            }
            for l in links
        ],
    })


@app.route("/api/network/links/add", methods=["POST"])
@admin_required
def api_network_links_add():
    """Add a network link (sector/link hardware).

    When the device is a MikroTik with IP + username + password, the router
    subscribers are pulled automatically into the customers table and pushed
    to the cloud — so the tower owner sees everything on any device instantly.
    """
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"ok": False, "error": "اسم الجهاز مطلوب"}), 400
    link_type = data.get("link_type", "MikroTik").strip()
    if link_type not in ("MikroTik", "Ubnt", "Mimosa"):
        link_type = "MikroTik"
    ip = data.get("ip", "").strip()
    if ip and not validate_host(ip):
        return jsonify({"ok": False, "error": "عنوان IP غير صالح"}), 400
    link_id = db.add_network_link(
        name=name, ip=ip, link_type=link_type,
        location=data.get("location", "").strip(),
        notes=data.get("notes", "").strip(),
        username=data.get("username", "").strip(),
        password=data.get("password", "").strip(),
        community=(data.get("community", "public").strip() or "public"),
    )
    _audit("اضافة جهاز شبكة", "network_link", link_id, f"تم اضافة الجهاز {name}")

    # ══ AUTO PULL: MikroTik device with credentials → pull + push instantly ══
    pull_summary = {}
    if link_type == "MikroTik" and ip and data.get("username", "").strip():
        try:
            from billing_system.mikrotik_sync import sync_pull_router, sync_push_cloud
            pull_summary = sync_pull_router(
                host=ip,
                username=data.get("username", "").strip(),
                password=data.get("password", ""),
                port=int(data.get("port", 8728) or 8728),
            )
            if pull_summary.get("secrets", 0) > 0:
                try:
                    sync_push_cloud()
                except Exception as e:  # noqa: BLE001
                    log.warning("[AutoSync] Cloud push failed: %s", e)
        except Exception as e:  # noqa: BLE001 — never break link creation
            log.error("[AutoSync] Pull failed for %s: %s", ip, e)
            pull_summary = {"errors": [str(e)]}

    return jsonify({"ok": True, "link_id": link_id, "auto_pull": pull_summary})


@app.route("/api/network/links/edit/<int:link_id>", methods=["POST"])
@admin_required
def api_network_links_edit(link_id):
    """Edit a network link.

    For a MikroTik device with IP + credentials, subscribers are pulled
    automatically so the tower owner's data stays current on every device.
    """
    link = db.get_network_link(link_id)
    if not link:
        return jsonify({"ok": False, "error": "الجهاز غير موجود"}), 404
    data = request.get_json() or {}
    ip = data.get("ip", "").strip()
    if ip and not validate_host(ip):
        return jsonify({"ok": False, "error": "عنوان IP غير صالح"}), 400
    db.update_network_link(
        link_id,
        name=data.get("name", link["name"]).strip(),
        ip=ip,
        link_type=data.get("link_type", link["link_type"]).strip(),
        location=data.get("location", "").strip(),
        notes=data.get("notes", "").strip(),
        username=data.get("username", "").strip(),
        password=data.get("password", "").strip(),
        community=(data.get("community", "public").strip() or "public"),
    )
    _audit("تعديل جهاز شبكة", "network_link", link_id, f"تم تعديل الجهاز {link['name']}")

    # ══ AUTO PULL: MikroTik device with credentials → pull + push instantly ══
    pull_summary = {}
    link_type = data.get("link_type", link["link_type"] or "").strip()
    if link_type == "MikroTik" and ip and data.get("username", "").strip():
        try:
            from billing_system.mikrotik_sync import sync_pull_router, sync_push_cloud
            pull_summary = sync_pull_router(
                host=ip,
                username=data.get("username", "").strip(),
                password=data.get("password", ""),
                port=int(data.get("port", 8728) or 8728),
            )
            if pull_summary.get("secrets", 0) > 0:
                try:
                    sync_push_cloud()
                except Exception as e:  # noqa: BLE001
                    log.warning("[AutoSync] Cloud push failed: %s", e)
        except Exception as e:  # noqa: BLE001 — never break link editing
            log.error("[AutoSync] Pull failed for %s: %s", ip, e)
            pull_summary = {"errors": [str(e)]}

    return jsonify({"ok": True, "auto_pull": pull_summary})


@app.route("/api/network/links/delete/<int:link_id>", methods=["POST"])
@admin_required
def api_network_links_delete(link_id):
    """Delete a network link."""
    link = db.get_network_link(link_id)
    if not link:
        return jsonify({"ok": False, "error": "الجهاز غير موجود"}), 404
    db.delete_network_link(link_id)
    _audit("حذف جهاز شبكة", "network_link", link_id, f"تم حذف الجهاز {link['name']}")
    return jsonify({"ok": True})


@app.route("/api/network/sync", methods=["POST"])
@admin_required
def api_network_sync():
    """Pull subscribers from every saved MikroTik sector and push to cloud.

    One-button sync: every stored network link of type MikroTik that has an
    IP + username is pulled (PPP secrets → customers) and the cloud snapshot
    is refreshed, so the tower owner sees the same data on any device.
    """
    from billing_system.mikrotik_sync import sync_pull_router, sync_push_cloud

    links = db.list_network_links()
    results = []
    total_secrets = 0
    errors = []
    for link in links:
        if (link.get("link_type") or "") != "MikroTik":
            continue
        ip = (link.get("ip") or "").strip()
        username = (link.get("username") or "").strip()
        if not ip or not username:
            continue
        try:
            pull = sync_pull_router(
                host=ip,
                username=username,
                password=link.get("password") or "",
                port=int(link.get("port") or 8728),
            )
            total_secrets += pull.get("secrets", 0)
            results.append({"name": link["name"], "ip": ip, **pull})
        except Exception as e:  # noqa: BLE001 — one failure must not block others
            log.error("[NetSync] %s (%s): %s", link.get("name"), ip, e)
            errors.append(f"{link.get('name', ip)}: {e}")
        break  # single primary MikroTik is the tower's PPPoE source

    if total_secrets > 0:
        try:
            push = sync_push_cloud()
        except Exception as e:  # noqa: BLE001
            log.warning("[NetSync] Cloud push failed: %s", e)
            push = {"ok": False, "error": str(e)}
    else:
        push = {}

    _audit("سحب مشتركين من الميكروتيك", "network", None,
           f"تم سحب {total_secrets} مشترك من الأجهزة")
    return jsonify({
        "ok": True,
        "links": results,
        "total_secrets": total_secrets,
        "errors": errors,
        "push": push,
    })


# ──────────────────────────────────────────────
#  PRINT ROUTES
# ──────────────────────────────────────────────

@app.route("/print/receipt/<int:payment_id>")
@login_required
def print_receipt(payment_id):
    pay = db.get_payment_details(payment_id)
    if not pay:
        return "الايصال غير موجود", 404
    invoice = db.get_invoice(pay["invoice_id"])
    remaining = (invoice["total_amount"] or 0) - (invoice["paid_amount"] or 0) if invoice else 0
    return render_template(
        "print_receipt.html", payment=pay, customer=pay, invoice=invoice,
        gen_info=get_gen_info(), remaining=remaining,
    )


@app.route("/print/a4/receipt/<int:payment_id>")
@login_required
def print_a4_receipt(payment_id):
    pay = db.get_payment_details(payment_id)
    if not pay:
        return "الايصال غير موجود", 404
    invoice = db.get_invoice(pay["invoice_id"])
    remaining = (invoice["total_amount"] or 0) - (invoice["paid_amount"] or 0) if invoice else 0
    return render_template(
        "print_a4_receipt.html", payment=pay, customer=pay, invoice=invoice,
        gen_info=get_gen_info(), remaining=remaining,
    )


@app.route("/print/invoice/<int:invoice_id>")
@login_required
def print_invoice(invoice_id):
    inv = db.get_invoice_with_customer(invoice_id)
    if not inv:
        return "الفاتورة غير موجودة", 404
    payments = db.list_invoice_payments(invoice_id)
    extras = db.list_invoice_extras(invoice_id)
    remaining = (inv["total_amount"] or 0) - (inv["paid_amount"] or 0)
    return render_template(
        "print_invoice_unified.html", invoice=inv, customer=inv,
        payments=payments, extras=extras, gen_info=get_gen_info(), remaining=remaining,
    )


@app.route("/print/a4/invoice/<int:invoice_id>")
@login_required
def print_a4_invoice(invoice_id):
    inv = db.get_invoice_with_customer(invoice_id)
    if not inv:
        return "الفاتورة غير موجودة", 404
    payments = db.list_invoice_payments(invoice_id)
    extras = db.list_invoice_extras(invoice_id)
    remaining = (inv["total_amount"] or 0) - (inv["paid_amount"] or 0)
    return render_template(
        "print_a4_invoice.html", invoice=inv, customer=inv,
        payments=payments, extras=extras, gen_info=get_gen_info(), remaining=remaining,
    )


# ──────────────────────────────────────────────
#  EXPORT
# ──────────────────────────────────────────────

def _style_helpers():
    """Return shared openpyxl style objects."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "header_fill": PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid"),
        "header_font": Font(name="Cairo", bold=True, color="FFFFFF", size=12),
        "body_font": Font(name="Cairo", size=11),
        "Alignment": Alignment,
    }


def _write_sheet(ws, title, headers, rows, styles):
    """Populate a worksheet with header + rows, right-to-left layout."""
    ws.title = title
    ws.views.sheetView[0].rightToLeft = True
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["Alignment"](horizontal="center", vertical="center")
        cell.border = styles["border"]
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = styles["body_font"]
            cell.alignment = styles["Alignment"](horizontal="right", vertical="center")
            cell.border = styles["border"]
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


@app.route("/api/export/excel")
@login_required
def api_export_excel():
    from openpyxl import Workbook

    customers = db.export_customers()
    invoices = db.export_invoices()
    payments = db.export_payments()

    wb = Workbook()
    styles = _style_helpers()
    status_map = {"active": "نشط", "inactive": "معطل", "suspended": "موقوف", "expired": "منتهي"}

    ws1 = wb.active
    _write_sheet(
        ws1, "المشتركين",
        ["ID", "الاسم", "الهاتف", "الهاتف 2", "العنوان", "المنطقة", "الباقة",
         "سعر الباقة", "اسم المستخدم", "IP", "نوع الجهاز", "الحالة",
         "تاريخ الاشتراك", "تاريخ التجديد", "ملاحظات"],
        [[
            c["id"], c["name"], c["phone"] or "", c["phone2"] or "", c["address"] or "",
            c["region"] or "", c["package_name"] or "", c["package_price"] or 0,
            c["username"] or "", c["ip_address"] or "", c["device_type"] or "",
            status_map.get(c["subscription_status"], c["subscription_status"]),
            fmt_dt(c["subscription_date"]), fmt_dt(c["renewal_date"]), c["notes"] or "",
        ] for c in customers]
    )

    ws2 = wb.create_sheet()
    _write_sheet(
        ws2, "الفواتير",
        ["ID", "المشترك", "الشهر", "السنة", "الباقة", "سعر الباقة",
         "الإجمالي", "المدفوع", "المتبقي", "الحالة"],
        [[
            inv["id"], inv["customer_name"], inv["month"], inv["year"], inv["package_name"] or "",
            int(round(float(inv["package_price"] or 0))),
            int(round(float(inv["total_amount"] or 0))),
            int(round(float(inv["paid_amount"] or 0))),
            int(round(float((inv["total_amount"] or 0) - (inv["paid_amount"] or 0)))),
            "مسدد" if inv["is_paid"] else ("مدفوع جزئياً" if inv["paid_amount"] > 0 else "غير مسدد"),
        ] for inv in invoices]
    )

    ws3 = wb.create_sheet()
    _write_sheet(
        ws3, "المدفوعات",
        ["ID", "المشترك", "المبلغ", "التاريخ", "طريقة الدفع", "ملاحظات"],
        [[
            p["id"], p["customer_name"], int(round(float(p["amount"]))),
            p["payment_date"], p["payment_method"], p["notes"] or "",
        ] for p in payments]
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    today = now_dt().strftime("%Y-%m-%d")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"isp_export_{today}.xlsx",
    )


@app.route("/api/export/payments/excel")
@login_required
def api_export_payments_excel():
    from openpyxl import Workbook

    payments = db.export_payments()
    wb = Workbook()
    ws = wb.active
    ws.title = "المدفوعات"
    ws.views.sheetView[0].rightToLeft = True
    styles = _style_helpers()

    headers = ["ID", "المشترك", "المبلغ", "التاريخ", "الوقت", "طريقة الدفع", "ملاحظات"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["Alignment"](horizontal="center", vertical="center")
        cell.border = styles["border"]

    for row_idx, p in enumerate(payments, 2):
        created_time = ""
        if p["created_at"]:
            try:
                created_time = datetime.strptime(str(p["created_at"]), DATETIME_DB_S).strftime("%I:%M %p")
            except ValueError:
                created_time = str(p["created_at"])
        row_data = [
            p["id"], p["customer_name"], int(round(float(p["amount"]))),
            p["payment_date"], created_time, p["payment_method"], p["notes"] or "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = styles["body_font"]
            cell.alignment = styles["Alignment"](horizontal="right", vertical="center")
            cell.border = styles["border"]

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"payments_export_{now_dt().strftime('%Y-%m-%d')}.xlsx",
    )


@app.route("/api/backup")
@admin_required
def api_backup():
    """Download the SQLite database file (admin only)."""
    if not os.path.exists(LOCAL_DB_PATH):
        return jsonify({"ok": False, "error": "قاعدة البيانات غير موجودة"}), 404
    _audit("نسخ احتياطي", "settings", None, "تم تحميل النسخة الاحتياطية لقاعدة البيانات")
    return send_file(
        LOCAL_DB_PATH,
        as_attachment=True,
        download_name=f"isp_backup_{now_dt().strftime('%Y-%m-%d')}.db",
        mimetype="application/x-sqlite3",
    )


# ──────────────────────────────────────────────
#  ENTRYPOINT
# ──────────────────────────────────────────────

if __name__ == "__main__":
    import threading
    import webbrowser

    # فتح المتصفح تلقائياً بعد بدء الخادم (حتى لا يبدو البرنامج "لا يعمل")
    threading.Timer(1.5, lambda: webbrowser.open("http://localhost:5000")).start()
    # debug=False مهم جداً للنسخة المجمعة (PyInstaller onefile):
    # debug=True يفعّل auto-reloader الذي يفرز عملية فرعية ويجعل
    # نافذة الـ EXE المجمع تختفي فوراً ولا تفتح شيئاً.
    app.run(debug=False, host="0.0.0.0", port=5000)
